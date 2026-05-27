import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(data: list, output_path: str = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "识别结果"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_font = Font(bold=True, size=11)
    cell_font = Font(size=11)

    col_widths = {}

    for row_idx, row_data in enumerate(data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical='center', wrap_text=True
            )
            if row_idx == 1:
                cell.font = header_font
            else:
                cell.font = cell_font

            col_letter = get_column_letter(col_idx)
            text_len = len(str(cell_value)) if cell_value else 0
            current_width = col_widths.get(col_letter, 0)
            col_widths[col_letter] = max(current_width, text_len)

    for col_letter, width in col_widths.items():
        adjusted = min(max(width * 1.5 + 2, 8), 50)
        ws.column_dimensions[col_letter].width = adjusted

    if output_path:
        wb.save(output_path)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
