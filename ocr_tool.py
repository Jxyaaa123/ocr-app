import sys
import os
import json
import sqlite3
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from PIL import Image, ImageTk
import numpy as np
import cv2
from rapidocr_onnxruntime import RapidOCR
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── OCR engine ──────────────────────────────────────────────

ocr_engine = RapidOCR()


def recognize(image_path: str) -> list:
    img = cv2.imread(image_path)
    if img is None:
        return [["无法读取图片"]]

    h, w = img.shape[:2]
    if max(h, w) > 3000:
        scale = 3000 / max(h, w)
        img = cv2.resize(img, (int(w*scale), int(h*scale)),
                         interpolation=cv2.INTER_AREA)

    result, _ = ocr_engine(img)
    if not result:
        return [["未识别到任何文字"]]

    items = []
    for line in result:
        box, text, conf = line
        box = np.array(box)
        cx = float(box[:, 0].mean())
        cy = float(box[:, 1].mean())
        y_top = float(box[:, 1].min())
        y_bot = float(box[:, 1].max())
        ih = y_bot - y_top
        items.append({'text': text, 'cx': cx, 'cy': cy,
                      'y_top': y_top, 'h': ih, 'conf': float(conf)})

    if not items:
        return [["未识别到任何文字"]]

    items.sort(key=lambda x: (x['y_top'], x['cx']))

    # group into rows by Y coordinate
    rows = []
    cur_row = [items[0]]

    for item in items[1:]:
        ref = cur_row[0]
        tol = max(ref['h'], item['h']) * 0.4
        if abs(item['y_top'] - ref['y_top']) < tol:
            cur_row.append(item)
        else:
            cur_row.sort(key=lambda x: x['cx'])
            rows.append(cur_row)
            cur_row = [item]
    cur_row.sort(key=lambda x: x['cx'])
    rows.append(cur_row)

    # assign columns: find column boundaries from all items' X positions
    all_cx = sorted(set(item['cx'] for row in rows for item in row))
    if len(all_cx) <= 1:
        return [[item['text'] for item in row] for row in rows]

    # cluster X positions into columns
    col_centers = []
    for cx in all_cx:
        merged = False
        for i, cc in enumerate(col_centers):
            if abs(cx - cc) < 40:
                col_centers[i] = (cc + cx) / 2
                merged = True
                break
        if not merged:
            col_centers.append(cx)
    col_centers.sort()

    # assign each item to nearest column
    result_data = []
    for row in rows:
        row_cells = [''] * len(col_centers)
        for item in row:
            best_col = 0
            best_dist = float('inf')
            for ci, cc in enumerate(col_centers):
                d = abs(item['cx'] - cc)
                if d < best_dist:
                    best_dist = d
                    best_col = ci
            row_cells[best_col] = item['text']
        result_data.append(row_cells)

    return result_data


# ── Excel export ────────────────────────────────────────────

def export_excel(data: list, path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "识别结果"

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    hdr_font = Font(bold=True, size=11)
    cell_font = Font(size=11)
    widths = {}

    for ri, row in enumerate(data, 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.font = hdr_font if ri == 1 else cell_font
            col = get_column_letter(ci)
            widths[col] = max(widths.get(col, 0), len(str(val or '')))

    for col, w in widths.items():
        ws.column_dimensions[col].width = min(max(w * 1.5 + 2, 8), 50)

    wb.save(path)


# ── History DB ──────────────────────────────────────────────

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "history.db")


def db_init():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""CREATE TABLE IF NOT EXISTS records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        image_path TEXT, data_json TEXT,
        row_count INTEGER, col_count INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    conn.commit()
    conn.close()


def db_save(image_path, data):
    conn = sqlite3.connect(DB_PATH)
    c = conn.execute(
        "INSERT INTO records (image_path,data_json,row_count,col_count) VALUES (?,?,?,?)",
        (image_path, json.dumps(data, ensure_ascii=False),
         len(data), max((len(r) for r in data), default=0)))
    rid = c.lastrowid
    conn.commit()
    conn.close()
    return rid


def db_list():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT id,row_count,col_count,created_at FROM records ORDER BY created_at DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def db_get(rid):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM records WHERE id=?", (rid,)).fetchone()
    conn.close()
    if row:
        d = dict(row)
        d['data'] = json.loads(d['data_json'])
        return d
    return None


def db_delete(rid):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM records WHERE id=?", (rid,))
    conn.commit()
    conn.close()


# ── GUI ─────────────────────────────────────────────────────

class App:
    def __init__(self):
        db_init()
        self.root = tk.Tk()
        self.root.title("图片表格识别")
        self.root.geometry("1000x700")
        self.root.minsize(800, 500)

        self.data = None
        self.image_path = None
        self._build_ui()

    def _build_ui(self):
        nb = ttk.Notebook(self.root)
        nb.pack(fill='both', expand=True, padx=8, pady=8)

        # ── tab 1: OCR ──
        tab1 = ttk.Frame(nb)
        nb.add(tab1, text="  识别导出  ")

        top = ttk.Frame(tab1)
        top.pack(fill='x', padx=8, pady=8)

        ttk.Button(top, text="选择图片", command=self._pick_image).pack(side='left')
        ttk.Button(top, text="开始识别", command=self._do_ocr).pack(side='left', padx=8)
        ttk.Button(top, text="下载 Excel", command=self._export).pack(side='left')
        ttk.Button(top, text="保存到历史", command=self._save_history).pack(side='left', padx=8)

        self.path_var = tk.StringVar(value="未选择图片")
        ttk.Label(top, textvariable=self.path_var).pack(side='left', padx=12)

        paned = ttk.PanedWindow(tab1, orient='horizontal')
        paned.pack(fill='both', expand=True, padx=8, pady=(0, 8))

        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=1)

        self.canvas = tk.Canvas(left_frame, bg='#f0f0f0')
        self.canvas.pack(fill='both', expand=True)

        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=2)

        cols = [f'C{i}' for i in range(20)]
        self.tree = ttk.Treeview(right_frame, columns=cols, show='headings',
                                 height=20)
        vsb = ttk.Scrollbar(right_frame, orient='vertical', command=self.tree.yview)
        hsb = ttk.Scrollbar(right_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        right_frame.columnconfigure(0, weight=1)
        right_frame.rowconfigure(0, weight=1)

        # ── tab 2: history ──
        tab2 = ttk.Frame(nb)
        nb.add(tab2, text="  历史记录  ")

        htop = ttk.Frame(tab2)
        htop.pack(fill='x', padx=8, pady=8)
        ttk.Button(htop, text="刷新", command=self._load_history).pack(side='left')

        hcols = ('id', 'rows', 'cols', 'time')
        self.hist_tree = ttk.Treeview(tab2, columns=hcols, show='headings', height=15)
        self.hist_tree.heading('id', text='ID')
        self.hist_tree.heading('rows', text='行数')
        self.hist_tree.heading('cols', text='列数')
        self.hist_tree.heading('time', text='时间')
        self.hist_tree.column('id', width=60)
        self.hist_tree.column('rows', width=60)
        self.hist_tree.column('cols', width=60)
        self.hist_tree.column('time', width=200)
        self.hist_tree.pack(fill='both', expand=True, padx=8, pady=(0, 8))

        hbtns = ttk.Frame(tab2)
        hbtns.pack(fill='x', padx=8, pady=(0, 8))
        ttk.Button(hbtns, text="查看", command=self._view_history).pack(side='left')
        ttk.Button(hbtns, text="导出 Excel", command=self._export_history).pack(
            side='left', padx=8)
        ttk.Button(hbtns, text="删除", command=self._delete_history).pack(side='left')

        self._load_history()

    def _pick_image(self):
        path = filedialog.askopenfilename(
            title="选择图片",
            filetypes=[("图片文件", "*.png *.jpg *.jpeg *.bmp *.tiff *.webp")])
        if not path:
            return
        self.image_path = path
        self.path_var.set(os.path.basename(path))

        img = Image.open(path)
        cw = self.canvas.winfo_width() or 400
        ch = self.canvas.winfo_height() or 400
        ratio = min(cw / img.width, ch / img.height, 1.0)
        new_size = (int(img.width * ratio), int(img.height * ratio))
        img = img.resize(new_size, Image.LANCZOS)
        self._photo = ImageTk.PhotoImage(img)
        self.canvas.delete('all')
        self.canvas.create_image(cw//2, ch//2, image=self._photo, anchor='center')

    def _do_ocr(self):
        if not self.image_path:
            messagebox.showwarning("提示", "请先选择图片")
            return
        self.root.config(cursor='watch')
        self.root.update()
        try:
            self.data = recognize(self.image_path)
            self._show_table(self.data)
        except Exception as e:
            messagebox.showerror("识别失败", str(e))
        finally:
            self.root.config(cursor='')

    def _show_table(self, data):
        for col in self.tree['columns']:
            self.tree.heading(col, text='')
        self.tree.delete(*self.tree.get_children())

        if not data:
            return

        max_cols = max(len(r) for r in data)
        for i in range(max_cols):
            self.tree.heading(f'C{i}', text=f'列{i+1}')

        for row in data:
            vals = row + [''] * (max_cols - len(row))
            self.tree.insert('', 'end', values=vals)

    def _export(self):
        if not self.data:
            messagebox.showwarning("提示", "请先识别图片")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="识别结果.xlsx")
        if path:
            export_excel(self.data, path)
            messagebox.showinfo("成功", f"已导出到 {path}")

    def _save_history(self):
        if not self.data:
            messagebox.showwarning("提示", "请先识别图片")
            return
        rid = db_save(self.image_path or "", self.data)
        messagebox.showinfo("成功", f"已保存 (ID: {rid})")
        self._load_history()

    def _load_history(self):
        self.hist_tree.delete(*self.hist_tree.get_children())
        for r in db_list():
            self.hist_tree.insert('', 'end', values=(
                r['id'], r['row_count'], r['col_count'], r['created_at']))

    def _get_selected_history_id(self):
        sel = self.hist_tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一条记录")
            return None
        return self.hist_tree.item(sel[0])['values'][0]

    def _view_history(self):
        rid = self._get_selected_history_id()
        if not rid:
            return
        rec = db_get(rid)
        if rec:
            self.data = rec['data']
            self._show_table(self.data)

    def _export_history(self):
        rid = self._get_selected_history_id()
        if not rid:
            return
        rec = db_get(rid)
        if not rec:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=f"记录_{rid}.xlsx")
        if path:
            export_excel(rec['data'], path)
            messagebox.showinfo("成功", f"已导出到 {path}")

    def _delete_history(self):
        rid = self._get_selected_history_id()
        if not rid:
            return
        if messagebox.askyesno("确认", "确定删除这条记录？"):
            db_delete(rid)
            self._load_history()

    def run(self):
        self.root.mainloop()


if __name__ == '__main__':
    App().run()
